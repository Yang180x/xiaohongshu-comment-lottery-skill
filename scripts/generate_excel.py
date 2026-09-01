#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""小红书寄样用户筛选 — Excel 报告生成器 v2.0

用法:
    python3 generate_excel.py <结果.json> [--out 输出目录] [--date YYYY-MM-DD]

输入: skill 阶段 7 产出的结果 JSON(结构见 SKILL.md「结果 JSON 数据结构」)。
输出: 小红书寄样用户筛选_YYYY-MM-DD.xlsx,4 个 Sheet:
      筛选总览 / 正式寄样名单 / 备选寄样名单 / 评论筛选明细
布局规范: 总览独立 Sheet(避免与表格列宽冲突);数据表统一
      表头样式+冻结+筛选+斑马纹+按内容自动行高+分列对齐。
缺失字段一律写「无法获取」,不做任何猜测。
"""

import argparse
import datetime
import json
import math
import os
import sys

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("错误: 缺少 openpyxl。安装: pip install openpyxl", file=sys.stderr)
    sys.exit(1)

NA = "无法获取"

# ---------------- 字段定义 ----------------
SHEET1_FIELDS = [
    "排名", "推荐等级", "用户昵称", "用户主页链接", "原始评论", "评论点赞数",
    "评论相关度", "产品兴趣度", "评论评分", "粉丝数", "关注数", "笔记数量",
    "获赞与收藏", "用户简介", "账号定位", "是否目标垂类", "是否作者粉丝",
    "推荐理由", "风险提示",
]
SHEET2_FIELDS = ["备选排名"] + [f for f in SHEET1_FIELDS if f != "排名"]
SHEET3_FIELDS = [
    "用户昵称", "用户主页链接", "评论内容", "评论点赞数", "内容相关度",
    "产品兴趣度", "需求匹配度", "评论真实性", "评论总评分",
    "是否进入主页分析", "是否最终入选", "未入选原因",
]

# 列宽(字段名 → 宽度)
COL_WIDTHS = {
    "排名": 6.5, "备选排名": 8.5, "推荐等级": 8.5, "用户昵称": 16,
    "用户主页链接": 30, "原始评论": 42, "评论内容": 42, "评论点赞数": 9,
    "评论相关度": 9.5, "产品兴趣度": 9.5, "评论评分": 8.5, "评论总评分": 9.5,
    "粉丝数": 8.5, "关注数": 8.5, "笔记数量": 8.5, "获赞与收藏": 10.5,
    "用户简介": 24, "账号定位": 13, "是否目标垂类": 10.5, "是否作者粉丝": 10.5,
    "推荐理由": 58, "风险提示": 18,
    "内容相关度": 9.5, "需求匹配度": 9.5, "评论真实性": 9.5,
    "是否进入主页分析": 12, "是否最终入选": 10.5, "未入选原因": 30,
}
# 长文本列(自动换行+左对齐+参与行高计算)
WRAP_FIELDS = {"原始评论", "评论内容", "推荐理由", "用户简介", "风险提示", "未入选原因"}
# 居中列(数字/短枚举)
CENTER_FIELDS = {
    "排名", "备选排名", "推荐等级", "评论点赞数", "评论相关度", "产品兴趣度",
    "评论评分", "评论总评分", "粉丝数", "关注数", "笔记数量", "获赞与收藏",
    "账号定位", "是否目标垂类", "是否作者粉丝", "内容相关度", "需求匹配度",
    "评论真实性", "是否进入主页分析", "是否最终入选",
}

# ---------------- 样式 ----------------
THIN = Side(style="thin", color="DDDDDD")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HEADER_FILL = PatternFill("solid", fgColor="534AB7")
HEADER_FONT = Font(name="PingFang SC", size=11, bold=True, color="FFFFFF")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
BODY_FONT = Font(name="PingFang SC", size=10.5, color="26215C")
LINK_FONT = Font(name="PingFang SC", size=10.5, color="185FA5", underline="single")
ZEBRA_FILL = PatternFill("solid", fgColor="F5F4FB")
LEVEL_STYLE = {
    "S": (PatternFill("solid", fgColor="F5A623"), Font(name="PingFang SC", size=10.5, bold=True, color="FFFFFF")),
    "A": (PatternFill("solid", fgColor="7BC86C"), Font(name="PingFang SC", size=10.5, bold=True, color="FFFFFF")),
    "B": (PatternFill("solid", fgColor="B4B2A9"), Font(name="PingFang SC", size=10.5, bold=True, color="FFFFFF")),
}
TITLE_FONT = Font(name="PingFang SC", size=16, bold=True, color="26215C")
SUBTITLE_FONT = Font(name="PingFang SC", size=11, color="5F5E5A")
LABEL_FONT = Font(name="PingFang SC", size=11, bold=True, color="3C3489")
VALUE_FONT = Font(name="PingFang SC", size=11, color="26215C")
LINE_FILL = PatternFill("solid", fgColor="534AB7")


def _val(row, key):
    v = row.get(key)
    if v is None or (isinstance(v, str) and v.strip() == ""):
        return NA
    return v


def _num(row, key):
    v = row.get(key)
    return v if isinstance(v, (int, float)) else 9999


def _text_lines(value, width):
    """估算换行后的行数(CJK 字符按 2 宽计)。"""
    if not isinstance(value, str) or not value:
        return 1
    total = 0
    for seg in value.split("\n"):
        w = sum(2 if ord(ch) > 127 else 1 for ch in seg)
        total += max(1, math.ceil(w / max(width - 2, 4)))
    return total


def _row_height(row, fields):
    lines = 1
    for f in fields:
        if f in WRAP_FIELDS:
            lines = max(lines, _text_lines(str(_val(row, f)), COL_WIDTHS.get(f, 12)))
    return max(22, min(220, lines * 15 + 8))


def _write_table(ws, fields, rows):
    """写入表头+数据,应用统一样式。返回表头行号(恒为 1)。"""
    # 表头
    for c_idx, field in enumerate(fields, 1):
        c = ws.cell(row=1, column=c_idx, value=field)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.border = BORDER
        c.alignment = HEADER_ALIGN
        ws.column_dimensions[get_column_letter(c_idx)].width = COL_WIDTHS.get(field, 12)
    ws.row_dimensions[1].height = 30

    # 数据
    for i, row in enumerate(rows):
        r_idx = i + 2
        ws.row_dimensions[r_idx].height = _row_height(row, fields)
        for c_idx, field in enumerate(fields, 1):
            value = _val(row, field)
            c = ws.cell(row=r_idx, column=c_idx, value=value)
            c.font = BODY_FONT
            c.border = BORDER
            # 斑马纹
            if i % 2 == 1:
                c.fill = ZEBRA_FILL
            # 链接列
            if field == "用户主页链接" and isinstance(value, str) and value.startswith("http"):
                c.hyperlink = value
                c.font = LINK_FONT
                c.alignment = Alignment(vertical="center", wrap_text=False)
                continue
            # 推荐等级徽章
            if field == "推荐等级" and value in LEVEL_STYLE:
                fill, font = LEVEL_STYLE[value]
                c.fill = fill
                c.font = font
                c.alignment = Alignment(horizontal="center", vertical="center")
                continue
            # 常规对齐
            if field in CENTER_FIELDS:
                c.alignment = Alignment(horizontal="center", vertical="center")
            elif field in WRAP_FIELDS:
                c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            else:
                c.alignment = Alignment(horizontal="left", vertical="center")

    ws.freeze_panes = "A2"
    if rows:
        last = f"{get_column_letter(len(fields))}{len(rows) + 1}"
        ws.auto_filter.ref = f"A1:{last}"
    else:
        ws.cell(row=2, column=1, value="(无数据)")
        ws.cell(row=2, column=1).font = SUBTITLE_FONT


def _write_overview(ws, data, date_str):
    """独立总览 Sheet:标题区 + 两栏信息卡。"""
    stats = data.get("statistics", {}) or {}
    note = data.get("note_summary", {}) or {}

    ws.column_dimensions["A"].width = 2.5
    for col in "BCDEFG":
        ws.column_dimensions[col].width = 15
    ws.column_dimensions["H"].width = 2.5

    # 标题
    ws.merge_cells("B2:G2")
    t = ws.cell(row=2, column=2, value="小红书寄样用户筛选报告")
    t.font = TITLE_FONT
    t.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 30

    ws.merge_cells("B3:G3")
    sub = ws.cell(row=3, column=3 - 1, value=f"生成日期:{date_str}    数据来源:小红书公开页面(BrowserSkill 采集)")
    sub.font = SUBTITLE_FONT
    ws.row_dimensions[3].height = 18

    # 分隔线
    for col in range(2, 8):
        ws.cell(row=4, column=col).fill = LINE_FILL
    ws.row_dimensions[4].height = 3

    # 左栏:笔记画像
    left = [
        ("笔记标题", _val(note, "标题")),
        ("作者", _val(note, "作者昵称")),
        ("产品或主题", _val(note, "产品或主题")),
        ("品类 / 主要卖点", f"{_val(note, '品类')}  |  {_val(note, '主要卖点')}"),
        ("目标人群", _val(note, "目标人群")),
        ("互动数据", f"赞 {_val(note, '点赞数')} · 藏 {_val(note, '收藏数')} · 评 {_val(note, '评论数')}"),
        ("笔记链接", data.get("note_url", NA)),
    ]
    # 右栏:筛选漏斗
    right = [
        ("目标 / 备选", f"{data.get('target_count', NA)} 正式 + {data.get('backup_count', NA)} 备选"),
        ("实际读取评论数", _val(stats, "实际读取评论数")),
        ("有效评论用户数", _val(stats, "有效评论用户数")),
        ("预筛候选数", _val(stats, "预筛候选数")),
        ("主页分析数", _val(stats, "主页分析数")),
        ("最终入选数", _val(stats, "最终入选数")),
    ]

    r = 6
    ws.merge_cells("B6:D6")
    h1 = ws.cell(row=6, column=2, value="笔记 / 产品画像")
    h1.font = Font(name="PingFang SC", size=12, bold=True, color="534AB7")
    ws.merge_cells("F6:G6")
    h2 = ws.cell(row=6, column=6, value="筛选漏斗")
    h2.font = Font(name="PingFang SC", size=12, bold=True, color="534AB7")

    r = 7
    max_r = r + max(len(left), len(right))
    for i in range(max(len(left), len(right))):
        rr = r + i
        ws.row_dimensions[rr].height = 20
        if i < len(left):
            label, value = left[i]
            ws.merge_cells(f"B{rr}:C{rr}")
            lc = ws.cell(row=rr, column=2, value=label)
            lc.font = LABEL_FONT
            lc.alignment = Alignment(horizontal="left", vertical="center")
            ws.merge_cells(f"D{rr}:E{rr}")
            vc = ws.cell(row=rr, column=4, value=value)
            vc.font = VALUE_FONT
            vc.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        if i < len(right):
            label, value = right[i]
            ws.merge_cells(f"F{rr}:F{rr}")
            lc = ws.cell(row=rr, column=6, value=label)
            lc.font = LABEL_FONT
            lc.alignment = Alignment(horizontal="left", vertical="center")
            ws.merge_cells(f"G{rr}:G{rr}")
            vc = ws.cell(row=rr, column=7, value=value)
            vc.font = VALUE_FONT
            vc.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # 说明
    note_line = _val(stats, "说明")
    if note_line != NA:
        rr = max_r + 1
        ws.merge_cells(f"B{rr}:G{rr}")
        nc = ws.cell(row=rr, column=2, value=f"说明:{note_line}")
        nc.font = SUBTITLE_FONT
        nc.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        ws.row_dimensions[rr].height = 34


def main():
    parser = argparse.ArgumentParser(description="小红书寄样用户筛选 Excel 生成器")
    parser.add_argument("json_file", help="结果 JSON 文件路径")
    parser.add_argument("--out", default=".", help="输出目录(默认当前目录)")
    parser.add_argument("--date", default=None, help="报告日期 YYYY-MM-DD(默认今天)")
    args = parser.parse_args()

    date_str = args.date or datetime.date.today().strftime("%Y-%m-%d")

    try:
        with open(args.json_file, "r", encoding="utf-8") as f:
            data = json.load(f)
    except (OSError, json.JSONDecodeError) as e:
        print(f"错误: 无法读取 JSON({e})", file=sys.stderr)
        sys.exit(1)

    finals = data.get("final_candidates", []) or []
    details = data.get("comment_details", []) or []
    official = sorted(
        [r for r in finals if r.get("名单类型", "正式") == "正式"],
        key=lambda r: _num(r, "排名"),
    )
    backup = sorted(
        [r for r in finals if r.get("名单类型") == "备选"],
        key=lambda r: _num(r, "备选排名") if isinstance(r.get("备选排名"), (int, float)) else _num(r, "排名"),
    )

    wb = Workbook()

    ws0 = wb.active
    ws0.title = "筛选总览"
    _write_overview(ws0, data, date_str)

    ws1 = wb.create_sheet("正式寄样名单")
    _write_table(ws1, SHEET1_FIELDS, official)

    ws2 = wb.create_sheet("备选寄样名单")
    _write_table(ws2, SHEET2_FIELDS, backup)

    ws3 = wb.create_sheet("评论筛选明细")
    _write_table(ws3, SHEET3_FIELDS, details)

    os.makedirs(args.out, exist_ok=True)
    out_path = os.path.join(args.out, f"小红书寄样用户筛选_{date_str}.xlsx")
    wb.save(out_path)

    print(f"已生成: {out_path}")
    print(f"  筛选总览 + 正式寄样名单 {len(official)} 人 + 备选寄样名单 {len(backup)} 人 + 评论筛选明细 {len(details)} 条")


if __name__ == "__main__":
    main()
